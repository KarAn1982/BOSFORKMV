from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).parent / "semantic_sources"
OUT = Path(__file__).parent / "semantic_analysis"
OUT.mkdir(exist_ok=True)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def read_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "utf-16"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8-replace"


def classify(query: str) -> list[str]:
    q = query.lower().replace("ё", "е")
    tags = []
    patterns = {
        "commercial": r"\b(купить|цена|стоимость|заказать|расчет|рассчитать|монтаж|установка|производитель)\b",
        "aluminum": r"\b(алюмин|alutech|алютех|alneo|алнео|schuco|шуко)\w*",
        "pvc": r"\b(пвх|пластиков|rehau|рехау)\w*",
        "windows": r"\b(окно|окна|оконн)\w*",
        "doors": r"\b(двер|входн\w* групп)\w*",
        "facades": r"\b(фасад|витраж|остеклен\w* фасад)\w*",
        "portals": r"\b(портал|раздвиж|панорам)\w*",
        "b2b": r"\b(застройщик|жк|жил\w* комплекс|генподряд|проектирован|тендер|объект)\w*",
        "service": r"\b(ремонт|сервис|регулиров|замен)\w*",
        "informational": r"\b(что|как|какой|виды|сравнение|отличие|характеристик|тепл\w* или холод)\w*",
        "geo": r"\b(пятигорск|ессентуки|кисловодск|ставропол|кмв|кавказ|сочи|краснодар|москва)\w*",
        "negative_risk": r"\b(своими руками|бесплатно|авито|б/у|бу|ваканси|работа|чертеж скачать|реферат)\b",
    }
    for tag, pattern in patterns.items():
        if re.search(pattern, q):
            tags.append(tag)
    return tags


sources = []
queries = []

for path in sorted(ROOT.glob("*.txt")):
    text, encoding = read_text(path)
    lines = [normalize(line) for line in text.splitlines() if normalize(line)]
    sources.append(
        {
            "file": path.name,
            "type": "txt",
            "encoding": encoding,
            "rows": len(lines),
            "sample": lines[:10],
        }
    )
    for line in lines:
        parts = re.split(r"[\t;]", line)
        query = normalize(parts[0])
        if query:
            queries.append(
                {
                    "query": query,
                    "source": path.name,
                    "sheet": "",
                    "frequency": next(
                        (
                            int(item.replace(" ", ""))
                            for item in parts[1:]
                            if re.fullmatch(r"[\d ]+", item.strip())
                        ),
                        None,
                    ),
                }
            )

for path in sorted(ROOT.glob("*.xlsx")):
    workbook = load_workbook(path, data_only=True, read_only=True)
    workbook_summary = {"file": path.name, "type": "xlsx", "sheets": []}
    for sheet in workbook.worksheets:
        rows = [
            [normalize(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        rows = [row for row in rows if any(row)]
        workbook_summary["sheets"].append(
            {
                "name": sheet.title,
                "rows": len(rows),
                "max_columns": max((len(row) for row in rows), default=0),
                "sample": rows[:5],
            }
        )
        if not rows:
            continue
        headers = [cell.lower() for cell in rows[0]]
        query_columns = [
            index
            for index, header in enumerate(headers)
            if any(
                token in header
                for token in (
                    "ключ",
                    "фраз",
                    "запрос",
                    "keyword",
                    "наименование",
                )
            )
        ]
        frequency_columns = [
            index
            for index, header in enumerate(headers)
            if any(token in header for token in ("частот", "показ", "прогноз"))
        ]
        if not query_columns:
            query_columns = [0]
        for row in rows[1:]:
            for column in query_columns[:2]:
                if column >= len(row):
                    continue
                query = normalize(row[column])
                if not query or len(query) < 3 or query.startswith("="):
                    continue
                frequency = None
                for freq_column in frequency_columns:
                    if freq_column < len(row):
                        match = re.search(r"\d+", row[freq_column].replace(" ", ""))
                        if match:
                            frequency = int(match.group())
                            break
                queries.append(
                    {
                        "query": query,
                        "source": path.name,
                        "sheet": sheet.title,
                        "frequency": frequency,
                    }
                )
    sources.append(workbook_summary)

deduplicated = {}
for row in queries:
    key = row["query"].lower().replace("ё", "е").strip(" \"'.,")
    if key not in deduplicated:
        deduplicated[key] = {
            "query": row["query"],
            "sources": [],
            "frequency": row["frequency"],
            "tags": classify(row["query"]),
        }
    item = deduplicated[key]
    source_ref = row["source"] + (f" / {row['sheet']}" if row["sheet"] else "")
    if source_ref not in item["sources"]:
        item["sources"].append(source_ref)
    if row["frequency"] is not None:
        item["frequency"] = max(item["frequency"] or 0, row["frequency"])

items = list(deduplicated.values())
tag_counts = Counter(tag for item in items for tag in item["tags"])
untagged = [item for item in items if not item["tags"]]

clusters = defaultdict(list)
for item in items:
    tags = set(item["tags"])
    if "negative_risk" in tags:
        cluster = "negative_or_irrelevant"
    elif "b2b" in tags or "facades" in tags:
        cluster = "b2b_facades_objects"
    elif "portals" in tags:
        cluster = "portals_panoramic"
    elif "doors" in tags and "aluminum" in tags:
        cluster = "aluminum_doors"
    elif "windows" in tags and "aluminum" in tags:
        cluster = "aluminum_windows"
    elif "windows" in tags and "pvc" in tags:
        cluster = "pvc_windows"
    elif "service" in tags:
        cluster = "service"
    elif "aluminum" in tags:
        cluster = "aluminum_general"
    elif "pvc" in tags:
        cluster = "pvc_general"
    else:
        cluster = "review_required"
    clusters[cluster].append(item)

result = {
    "sources": sources,
    "total_raw_queries": len(queries),
    "unique_queries": len(items),
    "tag_counts": dict(tag_counts),
    "cluster_counts": {key: len(value) for key, value in clusters.items()},
    "clusters": {
        key: sorted(
            value,
            key=lambda item: (
                -(item["frequency"] or 0),
                item["query"].lower(),
            ),
        )
        for key, value in clusters.items()
    },
    "untagged_count": len(untagged),
}

(OUT / "semantic_analysis.json").write_text(
    json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
)

lines = [
    "# Анализ семантических материалов BOSFOR",
    "",
    f"- Исходных строк: {len(queries)}",
    f"- Уникальных фраз: {len(items)}",
    f"- Требуют ручной классификации: {len(untagged)}",
    "",
    "## Кластеры",
]
for cluster, values in sorted(clusters.items(), key=lambda pair: -len(pair[1])):
    lines.append(f"- {cluster}: {len(values)}")

lines.extend(["", "## Источники"])
for source in sources:
    if source["type"] == "txt":
        lines.append(
            f"- {source['file']}: {source['rows']} строк, {source['encoding']}"
        )
    else:
        sheet_info = ", ".join(
            f"{sheet['name']} ({sheet['rows']})" for sheet in source["sheets"]
        )
        lines.append(f"- {source['file']}: {sheet_info}")

lines.extend(["", "## Верхние фразы по кластерам"])
for cluster, values in clusters.items():
    lines.append(f"\n### {cluster}")
    for item in sorted(
        values,
        key=lambda item: (-(item["frequency"] or 0), item["query"].lower()),
    )[:20]:
        frequency = f" — {item['frequency']}" if item["frequency"] else ""
        lines.append(f"- {item['query']}{frequency}")

(OUT / "summary.md").write_text("\n".join(lines), encoding="utf-8")
print("\n".join(lines[:40]))
